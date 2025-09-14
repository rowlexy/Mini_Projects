import openpyxl
# headers = ['OrderNum', 'Pickup Date', 'Origin', 'Destination', 'Status', 'ApptNum', 'Carrier', 'Delivery Date']
# wb = openpyxl.Workbook()
# ws = wb.active
# print(len(headers))
# for i, header in enumerate(headers, 1):
#     ws.cell(row=1, column=i, value=header)

# # wb.save('Load.xlsx')
wb = openpyxl.load_workbook('Test_Loads.xlsx')

ws = wb['Sheet1']

appt_numbers = ['0098345', '0098450', '0098572', '0098798', '0098682']
ws.cell(column=8, row=1, value='Slots Booked(Yes/No)')

for i, appt_num in enumerate(appt_numbers, 2):
    ws.cell(row=i+1, column=6, value=appt_num)
# print('Values successfully added')    
for row in range(2, ws.max_row + 1):
    appt_slot = ws[f'F{row}'].value
    if appt_slot and appt_slot != 'N/A':
        ws[f'H{row}'] = 'Yes'
    else:
        ws[f'H{row}'] = 'No'

ws.cell(row=1, column=9, value='Delivered (Yes/No)')
for row in range(2, ws.max_row + 1):
    delivery_status = ws[f'E{row}'].value
    if delivery_status and delivery_status == 'Delivered':
        ws.cell(row=row, column=9, value='Yes')
    else:
        ws.cell(row=row, column=9, value='No')
        
print('Values successfully added')    
wb.save('Test_Loads.xlsx')